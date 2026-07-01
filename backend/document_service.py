import io
import json
import logging
import os
import re
import tempfile
import zipfile
from copy import copy as shallow_copy
from datetime import datetime
from flask import session
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from config import Config
from onedrive_service import _headers, _handle_errors, _get_effective_path, _get_drive_base
import settings_store

logger = logging.getLogger(__name__)

def _cell_has_style(ws, row, col):
    """Check if a cell exists in the worksheet (was loaded from file) and is not a merged cell."""
    if (row, col) not in ws._cells:
        return False
    try:
        _ = ws._cells[(row, col)].value
        return True
    except AttributeError:
        return False


def _apply_cell_style(target, source):
    """Copy formatting from source cell to target cell."""
    if not source:
        return
    try:
        _ = source.has_style
    except AttributeError:
        return
    if not source.has_style:
        return
    try:
        target.font = shallow_copy(source.font)
        target.fill = shallow_copy(source.fill)
        target.border = shallow_copy(source.border)
        target.alignment = shallow_copy(source.alignment)
        target.number_format = source.number_format
    except Exception:
        pass


def _write_template_cell(ws, row, col, value, format_ref_row):
    """Write a cell value preserving existing formatting, or copying from reference row."""
    if (row, col) in ws._cells:
        cell = ws._cells[(row, col)]
        try:
            cell.value = value
        except AttributeError:
            pass
        return
    try:
        cell = ws.cell(row=row, column=col, value=value)
    except AttributeError:
        return
    if format_ref_row:
        src = ws.cell(row=format_ref_row, column=col)
        _apply_cell_style(cell, src)


def _clear_data_cells(ws, start_row, data_columns=None):
    """Clear values from data rows, preserving all cell formatting.
    Only clears cells in data_columns (set of column numbers) if provided.
    Skips merged cells."""
    max_col = ws.max_column if data_columns is None else max(data_columns)
    for r in range(start_row, ws.max_row + 1):
        for c in range(1, max_col + 1):
            if data_columns is not None and c not in data_columns:
                continue
            if (r, c) not in ws._cells:
                continue
            cell = ws._cells[(r, c)]
            try:
                cell.value = None
            except AttributeError:
                pass


def _normalize_date(val):
    """Convert various date formats to YYYY-MM-DD string. Returns None if unparseable."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if hasattr(val, 'strftime'):
        try:
            return val.strftime("%Y-%m-%d")
        except Exception:
            pass
    s = str(val).strip()
    if not s:
        return None
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        return s
    from datetime import date
    d = None
    for fmt in ("%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d",
                "%d/%m/%y", "%d-%b-%Y", "%d-%b-%y", "%d %B %Y"):
        try:
            d = datetime.strptime(s, fmt)
            break
        except ValueError:
            continue
    if d:
        return d.strftime("%Y-%m-%d")
    m = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", s)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return None


def _build_date_row_map(ws, date_col):
    """Build {normalized_date: row_number} from the template's date column."""
    date_row_map = {}
    for r in range(1, ws.max_row + 1):
        cell = ws.cell(row=r, column=date_col)
        norm = _normalize_date(cell.value)
        if norm:
            date_row_map[norm] = r
    return date_row_map


def _save_workbook(wb, keep_macro=None):
    """Save workbook to BytesIO, preserving VBA macros if present."""
    if keep_macro is None:
        keep_macro = hasattr(wb, 'vbaProject') and wb.vbaProject is not None
    suffix = ".xlsm" if keep_macro else ".xlsx"
    tmp = tempfile.NamedTemporaryFile(suffix=suffix, delete=False)
    try:
        tmp.close()
        wb.save(tmp.name)
        with open(tmp.name, "rb") as f:
            output = io.BytesIO(f.read())
        output.seek(0)
    finally:
        os.unlink(tmp.name)
    return output

GRAPH_BASE = "https://graph.microsoft.com/v1.0"


_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_XLSM_MIME = "application/vnd.ms-excel.sheet.macroEnabled.12"


def _xls_headers(filename):
    h = _headers()
    h["Content-Type"] = _XLSM_MIME if filename.lower().endswith(".xlsm") else _XLSX_MIME
    return h


COMBINED_HEADERS = [
    "Name",
    "Date",
    "Project ID",
    "Doc Task Type",
    "Doc ID",
    "Doc Version",
    "Doc Type",
    "Work Time",
    "Reviewer Time",
    "Doc Status",
    "Activity Code",
    "Activity Time",
    "Work Location",
    "Total Time",
]
HEADER_FILL = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
HEADER_FILL_GRAY = PatternFill(start_color="D1D5DB", end_color="D1D5DB", fill_type="solid")

# Two-row template column layout (row 1 = category, row 2 = field names)
TEMPLATE_2ROW_LAYOUT = [
    # (group_label, [(field_name, column_number), ...])
    ("Date", [("date", 1)]),
    ("Name", [("user_name", 2)]),
    ("Project Document Work #1", [
        ("project_id", 3), ("doc_task_type", 4), ("doc_id", 5),
        ("doc_version", 6), ("doc_type", 7), ("work_time", 8),
        ("reviewer_time", 9), ("doc_status", 10),
    ]),
    ("Project Document Work #2", [
        ("project_id", 11), ("doc_task_type", 12), ("doc_id", 13),
        ("doc_version", 14), ("doc_type", 15), ("work_time", 16),
        ("reviewer_time", 17), ("doc_status", 18),
    ]),
    ("Activity #1", [
        ("activity_code", 19), ("project_id", 20), ("activity_time", 21),
    ]),
    ("Activity #2", [
        ("activity_code", 22), ("project_id", 23), ("activity_time", 24),
    ]),
    ("Activity #3", [
        ("activity_code", 25), ("project_id", 26), ("activity_time", 27),
    ]),
    ("Summary", [
        ("work_location", 28), ("leave_travel", 29),
        ("act_time_sum", 30), ("total_hours", 31), ("remarks", 32),
    ]),
]

TEMPLATE_2ROW_ROW1_LABELS = {
    1: "Name",
    2: "Date",
    3: "Project Document Work (Prepare, Check, Review, Mentoring)-#1",
    11: "Project Document Work (Prepare, Check, Review, Mentoring)-#2",
    19: "Other Project Activity or Non-Revenue Activity",
    28: "Leave/Travel Time",
    30: "Total",
    32: "Remarks",
}

HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _ensure_documents_folder(drive_base=None):
    if drive_base is None:
        drive_base = _get_drive_base()
    path = _get_effective_path().strip("/")
    if not path:
        return "root"

    resp = requests.get(
        f"{GRAPH_BASE}{drive_base}/root:/{path}",
        headers=_headers(),
    )
    if resp.ok:
        return resp.json().get("id")

    current = "root"
    for part in path.split("/"):
        create_resp = requests.post(
            f"{GRAPH_BASE}{drive_base}/items/{current}/children",
            headers=_headers(),
            json={"name": part, "folder": {}, "@microsoft.graph.conflictBehavior": "fail"},
        )
        if create_resp.status_code in (200, 201):
            current = create_resp.json().get("id")
        else:
            check = requests.get(
                f"{GRAPH_BASE}{drive_base}/items/{current}:/{part}",
                headers=_headers(),
            )
            if check.ok:
                current = check.json().get("id")
            else:
                return "root"
    return current


def _excel_filename(user_email, is_macro=False):
    safe = user_email.replace("@", "_at_").replace(".", "_dot_")
    ext = ".xlsm" if is_macro else ".xlsx"
    return f"timesheet_{safe}{ext}"


DOC_HEADER_MAP = {
    "user_name": ["name", "Name", "User Name", "user name"],
    "date": ["date", "Date"],
    "project_id": ["project id", "project", "Project ID", "Project Id"],
    "doc_task_type": ["task type", "document task type", "doc task type", "Task Type", "Doc Task Type"],
    "doc_id": ["doc id", "document id", "doc Id", "Doc ID", "Document ID"],
    "doc_version": ["doc version", "target doc version", "drs", "version", "Doc Version", "Version"],
    "doc_type": ["doc type", "document type", "Doc Type", "Document Type"],
    "work_time": ["work time", "work time on document", "Work Time"],
    "reviewer_time": ["reviewer time", "reviewer's / mentor's time", "mentor time", "Reviewer Time"],
    "doc_status": ["doc status", "doc. status", "document status", "status", "Doc Status"],
    "work_location": ["work location", "location", "Work Location"],
}

OTHER_HEADER_MAP = {
    "date": ["date", "Date"],
    "activity_code": ["activity code", "Activity Code"],
    "project_id": ["project id", "Project ID"],
    "activity_time": ["activity time", "Activity Time"],
    "work_location": ["work location", "location", "Work Location"],
}


def _has_two_row_header(ws):
    if ws.max_row < 2:
        logger.debug("_has_two_row_header: max_row < 2")
        return False
    max_c = ws.max_column or 1
    if max_c < 40:
        max_c = 40
    keywords = ["Project Document Work", "Other Project Activity",
                "Leave/Travel", "Total", "Remarks"]
    for c in range(1, max_c + 1):
        val = str(ws.cell(row=1, column=c).value or "").strip()
        if val:
            for kw in keywords:
                if kw.lower() in val.lower():
                    logger.debug("_has_two_row_header: row1 col%d match '%s' in '%s'", c, kw, val)
                    return True
    logger.debug("_has_two_row_header: no row1 match, checking row2")
    row2_field_keywords = ["activity code", "project id", "activity time",
                          "doc task type", "doc id", "doc version",
                          "doc type", "work time", "reviewer time",
                          "doc status", "work location"]
    for c in range(1, max_c + 1):
        val = str(ws.cell(row=2, column=c).value or "").strip()
        if val:
            logger.debug("_has_two_row_header: row2 col%d='%s'", c, val)
            for kw in row2_field_keywords:
                if kw in val.lower():
                    logger.debug("_has_two_row_header: row2 col%d match '%s' in '%s'", c, kw, val)
                    return True
    logger.debug("_has_two_row_header: no match in row2 either")
    return False


def _scan_row(ws, row_num, field_map, multi=False):
    col_map = {} if not multi else {}
    max_c = ws.max_column or 1
    if max_c < 40:
        max_c = 40
    for c in range(1, max_c + 1):
        cell = ws.cell(row=row_num, column=c)
        raw = cell.value
        if raw:
            val = str(raw).strip().lower()
            for field, candidates in field_map.items():
                for cand in candidates:
                    cv = cand.lower()
                    if field == "date":
                        if re.search(r"(?<!\w)" + re.escape(cv) + r"(?!\w)", val):
                            if multi:
                                col_map.setdefault(field, []).append(c)
                            else:
                                col_map[field] = c
                            break
                    elif cv in val:
                        if multi:
                            col_map.setdefault(field, []).append(c)
                        else:
                            col_map[field] = c
                        break
    return col_map


def _build_column_map_two_row(ws, field_map):
    return _scan_row(ws, 2, field_map, multi=True)


def _build_column_map(ws, field_map):
    return _scan_row(ws, 1, field_map, multi=False)


HEADER_TO_FIELD = {}
for field, candidates in DOC_HEADER_MAP.items():
    for c in candidates:
        HEADER_TO_FIELD[c.lower()] = field


def _ensure_doc_headers(ws, doc_map):
    next_col = ws.max_column + 1 if ws.max_column else 1
    def style_cell(c):
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGNMENT
        c.border = THIN_BORDER
    existing_headers = set(doc_map.keys())
    for header in DEFAULT_HEADERS:
        field = HEADER_TO_FIELD.get(header.lower())
        if field and field not in existing_headers:
            cell = ws.cell(row=1, column=next_col, value=header)
            style_cell(cell)
            doc_map[field] = next_col
            next_col += 1
    return doc_map


OTHER_HEADER_TO_FIELD = {}
for field, candidates in OTHER_HEADER_MAP.items():
    for c in candidates:
        OTHER_HEADER_TO_FIELD[c.lower()] = field


def _ensure_other_headers(ws, other_map):
    OTHER_HEADERS = ["Date", "Activity Code", "Project ID", "Activity Time", "Work Location"]
    next_col = ws.max_column + 1 if ws.max_column else 1
    def style_cell(c):
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = HEADER_ALIGNMENT
        c.border = THIN_BORDER
    existing_headers = set(other_map.keys())
    for header in OTHER_HEADERS:
        field = OTHER_HEADER_TO_FIELD.get(header.lower())
        if field and field not in existing_headers:
            cell = ws.cell(row=1, column=next_col, value=header)
            style_cell(cell)
            other_map[field] = next_col
            next_col += 1
    return other_map


ALL_HEADER_MAP = {}
for field, candidates in DOC_HEADER_MAP.items():
    ALL_HEADER_MAP[field] = list(candidates)
for field, candidates in OTHER_HEADER_MAP.items():
    if field in ALL_HEADER_MAP:
        seen = set(ALL_HEADER_MAP[field])
        for c in candidates:
            if c not in seen:
                ALL_HEADER_MAP[field].append(c)
                seen.add(c)
    else:
        ALL_HEADER_MAP[field] = list(candidates)

ALL_HEADER_MAP.setdefault("total_time", ["Total Time", "total time", "Total", "total"])
ALL_HEADER_MAP.setdefault("total_hours", ["Total Hours", "total hours"])
ALL_HEADER_MAP.setdefault("remarks", ["Remarks"])
ALL_HEADER_MAP.setdefault("leave_travel", ["Leave/Travel"])
ALL_HEADER_MAP.setdefault("act_summary_time", ["Activity Time Total", "activity time total", "Total Activity Time"])

TEMPLATE_2ROW_EXTRA = {
    "user_name": ["Name", "User Name"],
    "project_id": ["Project ID #1", "Project ID #2", "Project ID (If Proj. Activity)"],
    "doc_task_type": ["Document Task Type (P= Prepare or C=Check)"],
    "doc_id": ["Doc ID"],
    "doc_version": ["Target Doc Version / DRS for doc. Version"],
    "doc_type": ["Doc Type N=New Doc U=Update"],
    "work_time": ["Work time on Document"],
    "reviewer_time": ["Reviewer's / Mentor's time"],
    "doc_status": ["Doc. Status"],
    "activity_code": ["Activity Code #1", "Activity Code #2", "Activity Code #3"],
    "activity_time": ["Activity Time"],
    "work_location": ["Work location"],
    "date": ["dd/mm/yy"],
}
for field, variants in TEMPLATE_2ROW_EXTRA.items():
    if field in ALL_HEADER_MAP:
        seen = set(c.lower() for c in ALL_HEADER_MAP[field])
        for v in variants:
            if v.lower() not in seen:
                ALL_HEADER_MAP[field].append(v)
                seen.add(v.lower())


def _write_positional_fallback(ws, entries):
    date_row_map = _build_date_row_map(ws, 1)
    REF_ROW = 2 if ws.max_row >= 2 else 1

    grouped = {}
    for e in entries:
        d = _normalize_date(e.get("date", ""))
        if not d:
            continue
        if d not in grouped:
            grouped[d] = []
        grouped[d].append(e)

    for date_str, day_entries in grouped.items():
        row = date_row_map.get(date_str)
        if row is None:
            continue
        doc_entries = [e for e in day_entries if e.get("activity_type") not in ("other", "leave_travel")]
        other_entries = [e for e in day_entries if e.get("activity_type") == "other"]
        leave_entries = [e for e in day_entries if e.get("activity_type") == "leave_travel"]

        doc_fields = ["project_id", "doc_task_type", "doc_id", "doc_version",
                      "doc_type", "work_time", "reviewer_time", "doc_status"]

        for slot_i in range(2):
            entry = doc_entries[slot_i] if slot_i < len(doc_entries) else None
            base_col = 2 + slot_i * 8
            if entry:
                for fi, field in enumerate(doc_fields):
                    val = entry.get(field)
                    if val:
                        _write_template_cell(ws, row, base_col + fi, val, REF_ROW)

        for slot_i in range(3):
            entry = other_entries[slot_i] if slot_i < len(other_entries) else None
            if entry:
                code = entry.get("activity_code")
                proj = entry.get("project_id")
                tm = entry.get("activity_time")
                if slot_i == 0:
                    ws.cell(row=row, column=18).value = code
                    ws.cell(row=row, column=19).value = proj
                    c = ws.cell(row=row, column=20)
                    c.value = tm / 24.0 if tm else None
                    c.number_format = 'h:mm'
                elif slot_i == 1:
                    ws.cell(row=row, column=21).value = code
                    ws.cell(row=row, column=22).value = proj
                    c = ws.cell(row=row, column=23)
                    c.value = tm / 24.0 if tm else None
                    c.number_format = 'h:mm'
                elif slot_i == 2:
                    ws.cell(row=row, column=24).value = code
                    ws.cell(row=row, column=25).value = proj

        locations = set()
        for e in day_entries:
            loc = e.get("work_location", "")
            if loc:
                locations.add(loc)
        if locations:
            _write_template_cell(ws, row, 27, ", ".join(sorted(locations)), REF_ROW)

        remarks_list = [e.get("remarks", "") for e in doc_entries + other_entries if e.get("remarks")]
        if remarks_list:
            _write_template_cell(ws, row, 31, "; ".join(remarks_list), REF_ROW)

        total_work = sum(float(e.get("work_time", 0) or 0) for e in doc_entries)
        total_review = sum(float(e.get("reviewer_time", 0) or 0) for e in doc_entries)
        total_activity = sum(float(e.get("activity_time", 0) or 0) for e in other_entries)
        total_leave = sum(float(e.get("time", 0) or 0) for e in leave_entries)
        grand_total = total_work + total_review + total_activity + total_leave

        total = total_work + total_review + total_activity
        if total:
            c = ws.cell(row=row, column=30)
            c.value = total / 24.0
            c.number_format = 'h:mm'

        if total_leave:
            c = ws.cell(row=row, column=29)
            c.value = total_leave / 24.0
            c.number_format = 'h:mm'


def _build_from_template(entries, template_bytes):
    try:
        from io import BytesIO
        buf = BytesIO(template_bytes)
        wb = load_workbook(buf, keep_vba=True)
        _preserve_vba = hasattr(wb, 'vbaProject') and wb.vbaProject is not None
    except Exception:
        return _build_from_scratch(entries)
    ws = wb.active

    if _has_two_row_header(ws):
        row2_map = _build_column_map_two_row(ws, ALL_HEADER_MAP)
        if row2_map.get("activity_code") or row2_map.get("activity_time"):
            return _build_from_template_2row(entries, wb)

    col_map = _build_column_map(ws, ALL_HEADER_MAP)
    if col_map and any(f in col_map for f in ("activity_code", "project_id", "activity_time")):
        return _write_single_row(ws, entries, col_map, 1)

    row2_map = _build_column_map_two_row(ws, ALL_HEADER_MAP)
    if row2_map.get("activity_code") or row2_map.get("activity_time"):
        return _build_from_template_2row(entries, wb)

    _write_positional_fallback(ws, entries)
    output = _save_workbook(wb)
    return output, _preserve_vba


def _write_single_row(ws, entries, col_map, header_row):
    date_col = col_map.get("date")
    if date_col:
        date_row_map = _build_date_row_map(ws, date_col)
    else:
        date_row_map = {}
    for entry in entries:
        entry_date = entry.get("date", "")
        if not entry_date:
            continue
        norm_date = _normalize_date(entry_date)
        if not norm_date:
            continue
        row = date_row_map.get(norm_date)
        if row is None:
            continue
        try:
            for field, col in col_map.items():
                if field in ("date", "user_name"):
                    continue
                if field == "total_time":
                    wt = float(entry.get("work_time", 0) or 0)
                    rt = float(entry.get("reviewer_time", 0) or 0)
                    at = float(entry.get("activity_time", 0) or 0)
                    val = wt + rt + at
                    if val:
                        _write_template_cell(ws, row, col, val, header_row)
                else:
                    val = entry.get(field)
                    if val:
                        _write_template_cell(ws, row, col, val, header_row)
        except Exception as e:
            logger.warning("Error writing entry %s to row %d: %s", entry.get("id", ""), row, e)
            continue
    return _save_workbook(ws.parent), False


def _build_from_template_2row(entries, wb):
    ws = wb.active
    HEADER_ROW = 2
    DATA_START = 3

    col_map = _build_column_map_two_row(ws, ALL_HEADER_MAP)
    row1_map = _build_column_map(ws, ALL_HEADER_MAP)
    for field, col in row1_map.items():
        if field not in col_map or not col_map[field]:
            col_map[field] = [col]
        elif col not in col_map[field]:
            col_map[field].append(col)
    if not col_map:
        _write_positional_fallback(ws, entries)
        output = _save_workbook(wb)
        return output, False

    date_col = None
    if "date" in col_map and col_map["date"]:
        date_col = col_map["date"][0]
        date_row_map = _build_date_row_map(ws, date_col)
    else:
        date_row_map = {}

    def get_col(field, slot_index):
        lst = col_map.get(field, [])
        if slot_index < len(lst):
            return lst[slot_index]
        return None

    grouped = {}
    for e in entries:
        d = _normalize_date(e.get("date", ""))
        if not d:
            continue
        if d not in grouped:
            grouped[d] = []
        grouped[d].append(e)

    for date_str, day_entries in grouped.items():
        row = date_row_map.get(date_str)
        if row is None:
            continue
        doc_entries = [e for e in day_entries if e.get("activity_type") not in ("other", "leave_travel")]
        other_entries = [e for e in day_entries if e.get("activity_type") == "other"]
        leave_entries = [e for e in day_entries if e.get("activity_type") == "leave_travel"]

        if leave_entries:
            leave_val = ", ".join(e.get("leave_travel_type", "") for e in leave_entries)
            if leave_val:
                if "leave_travel" in col_map and col_map["leave_travel"]:
                    _write_template_cell(ws, row, col_map["leave_travel"][0], leave_val, HEADER_ROW)

        doc_fields = ["project_id", "doc_task_type", "doc_id", "doc_version",
                      "doc_type", "work_time", "reviewer_time", "doc_status"]

        for slot_i in range(2):
            entry = doc_entries[slot_i] if slot_i < len(doc_entries) else None
            for field in doc_fields:
                col = get_col(field, slot_i)
                if col is None:
                    continue
                if entry:
                    val = entry.get(field)
                    if val:
                        _write_template_cell(ws, row, col, val, HEADER_ROW)

        remarks_list = [e.get("remarks", "") for e in doc_entries + other_entries if e.get("remarks")]
        if remarks_list:
            _write_template_cell(ws, row, 31, "; ".join(remarks_list), HEADER_ROW)

        act_fields = ["activity_code", "project_id", "activity_time"]

        for slot_i in range(3):
            entry = other_entries[slot_i] if slot_i < len(other_entries) else None
            for field in act_fields:
                if field == "project_id":
                    act_project_ids = [c for c in col_map.get("project_id", []) if c >= 19]
                    col = act_project_ids[slot_i] if slot_i < len(act_project_ids) else None
                else:
                    col = get_col(field, slot_i)
                if col is None:
                    continue
                if entry:
                    val = entry.get(field)
                    if val:
                        _write_template_cell(ws, row, col, val, HEADER_ROW)

        locations = set()
        for e in day_entries:
            loc = e.get("work_location", "")
            if loc:
                locations.add(loc)
        if locations:
            if "work_location" in col_map and col_map["work_location"]:
                _write_template_cell(ws, row, col_map["work_location"][0], ", ".join(sorted(locations)), HEADER_ROW)

        total_work = sum(float(e.get("work_time", 0) or 0) for e in doc_entries)
        total_review = sum(float(e.get("reviewer_time", 0) or 0) for e in doc_entries)
        total_activity = sum(float(e.get("activity_time", 0) or 0) for e in other_entries)
        total_leave = sum(float(e.get("time", 0) or 0) for e in leave_entries)
        grand_total = total_work + total_review + total_activity + total_leave

        if grand_total:
            if "total_hours" in col_map and col_map["total_hours"]:
                _write_template_cell(ws, row, col_map["total_hours"][0], grand_total, HEADER_ROW)

        if total_activity:
            if "act_summary_time" in col_map and col_map["act_summary_time"]:
                _write_template_cell(ws, row, col_map["act_summary_time"][0], total_activity, HEADER_ROW)

    output = _save_workbook(wb)
    return output, False


def _write_scratch_headers(ws):
    row1_groups = [
        (1, 1, "Date"),
        (2, 2, "Name"),
        (3, 10, "Project Document Work (Prepare, Check, Review, Mentoring)-#1"),
        (11, 18, "Project Document Work (Prepare, Check, Review, Mentoring)-#2"),
        (19, 27, "Other Project Activity or Non-Revenue Activity"),
        (28, 30, "Leave/Travel Time"),
        (31, 31, "Total"),
        (32, 32, "Remarks"),
    ]
    for start, end, label in row1_groups:
        if start < end:
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        cell = ws.cell(row=1, column=start, value=label)
        cell.font = Font(bold=True, size=10)
        cell.fill = HEADER_FILL_GRAY
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        if start < end:
            for c in range(start + 1, end + 1):
                ws.cell(row=1, column=c).border = THIN_BORDER

    row2_headers = [
        ("Date", 1),
        ("Name", 2),
        ("Project ID #1", 3),
        ("Document Task Type (P= Prepare or C=Check)", 4),
        ("Doc ID", 5),
        ("Target Doc Version / DRS for doc. Version", 6),
        ("Doc Type N=New Doc U=Update", 7),
        ("Work time on Document", 8),
        ("Reviewer's / Mentor's time", 9),
        ("Doc. Status", 10),
        ("Project ID #2", 11),
        ("Document Task Type (P= Prepare or C=Check)", 12),
        ("Doc ID", 13),
        ("Target Doc Version / DRS for doc. Version", 14),
        ("Doc Type N=New Doc U=Update", 15),
        ("Work time on Document", 16),
        ("Reviewer's / Mentor's time", 17),
        ("Doc. Status", 18),
        ("Activity Code #1", 19),
        ("Project ID (If Proj. Activity)", 20),
        ("Activity Time", 21),
        ("Activity Code #2", 22),
        ("Project ID (If Proj. Activity)", 23),
        ("Activity Time", 24),
        ("Activity Code #3", 25),
        ("Project ID (If Proj. Activity)", 26),
        ("Activity Time", 27),
        ("Work location", 28),
        ("Leave/Travel", 29),
        ("Activity Time", 30),
        ("Total Hours", 31),
        ("Remarks", 32),
    ]

    for header, col in row2_headers:
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    widths = [14, 20, 14, 16, 22, 14, 12, 12, 14, 14, 14, 16, 22, 14, 12, 12, 14, 14, 18, 18, 14, 18, 18, 14, 18, 18, 14, 14, 14, 14, 14, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _build_from_scratch(entries):
    wb = Workbook()
    ws = wb.active
    ws.title = "Timesheet"
    _write_scratch_headers(ws)

    grouped = {}
    for e in entries:
        d = e.get("date", "")
        if d not in grouped:
            grouped[d] = []
        grouped[d].append(e)

    sorted_dates = sorted(grouped.keys())
    row = 3

    for date_str in sorted_dates:
        day_entries = grouped[date_str]
        doc_entries = [e for e in day_entries if e.get("activity_type") not in ("other", "leave_travel")]
        other_entries = [e for e in day_entries if e.get("activity_type") == "other"]
        leave_entries = [e for e in day_entries if e.get("activity_type") == "leave_travel"]

        user_name = next((e.get("user_name", "") for e in day_entries if e.get("user_name")), "")

        def set_cell(col, val):
            c = ws.cell(row=row, column=col, value=val)
            c.border = THIN_BORDER
            c.alignment = Alignment(vertical="center")

        set_cell(1, date_str)
        set_cell(2, user_name)

        doc_fields = ["project_id", "doc_task_type", "doc_id", "doc_version",
                      "doc_type", "work_time", "reviewer_time", "doc_status"]

        for slot_i in range(2):
            entry = doc_entries[slot_i] if slot_i < len(doc_entries) else None
            base_col = 2 + slot_i * 8
            if entry:
                for fi, field in enumerate(doc_fields):
                    val = entry.get(field, "")
                    if val is None:
                        val = ""
                    set_cell(base_col + fi, val)
            else:
                for fi in range(8):
                    set_cell(base_col + fi, "")

        act_fields = ["activity_code", "project_id", "activity_time"]
        slot_sizes = [3, 3, 2]

        for slot_i in range(3):
            entry = other_entries[slot_i] if slot_i < len(other_entries) else None
            base_col = 19 + slot_i * 3
            if entry:
                for fi, field in enumerate(act_fields):
                    val = entry.get(field, "")
                    if val is None:
                        val = ""
                    set_cell(base_col + fi, val)
            else:
                for fi in range(3):
                    set_cell(base_col + fi, "")

        locations = set()
        for e in day_entries:
            loc = e.get("work_location", "")
            if loc:
                locations.add(loc)
        work_location = ", ".join(sorted(locations)) if locations else ""

        total_work = sum(float(e.get("work_time", 0) or 0) for e in doc_entries)
        total_review = sum(float(e.get("reviewer_time", 0) or 0) for e in doc_entries)
        total_activity = sum(float(e.get("activity_time", 0) or 0) for e in other_entries)
        total_leave = sum(float(e.get("time", 0) or 0) for e in leave_entries)
        grand_total = total_work + total_review + total_activity + total_leave

        leave_val = ", ".join(e.get("leave_travel_type", "") for e in leave_entries) if leave_entries else ""

        remarks_list = [e.get("remarks", "") for e in doc_entries + other_entries if e.get("remarks")]
        remarks_val = "; ".join(remarks_list) if remarks_list else ""

        set_cell(28, work_location)
        set_cell(29, leave_val)
        set_cell(30, total_activity)
        set_cell(31, remarks_val)
        set_cell(32, grand_total)

        row += 1

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, False


def _build_workbook(entries, template_bytes=None):
    if template_bytes:
        return _build_from_template(entries, template_bytes)
    return _build_from_scratch(entries)


def _get_template_filename():
    user = session.get("user", {})
    email = user.get("email", "unknown")
    name = settings_store.get_template_filename(email)
    return name if name else None


def generate_excel(entries, upload_template_bytes=None, target_filename=None):
    user = session.get("user", {})
    email = user.get("email", "unknown")
    drive_base = _get_drive_base()
    folder_id = _ensure_documents_folder(drive_base=drive_base) or "root"

    workbook_bytes, is_macro = _build_workbook(entries, upload_template_bytes)
    if not target_filename:
        target_filename = _get_template_filename()
    filename = target_filename or _excel_filename(email, is_macro)
    content = workbook_bytes.read()

    def _try_put(fb_id):
        folder_children = requests.get(
            f"{GRAPH_BASE}{drive_base}/items/{fb_id}/children",
            headers=_headers(),
        )
        existing_id = None
        if folder_children.ok:
            for child in folder_children.json().get("value", []):
                if child.get("name") == filename:
                    existing_id = child["id"]
                    break
        if existing_id:
            return requests.put(
                f"{GRAPH_BASE}{drive_base}/items/{existing_id}/content",
                headers=_xls_headers(filename), data=content,
            )
        return requests.put(
            f"{GRAPH_BASE}{drive_base}/items/{fb_id}:/{filename}:/content",
            headers=_xls_headers(filename), data=content,
        )

    resp = _try_put(folder_id)
    if _handle_errors(resp):
        return generate_excel(entries, upload_template_bytes, target_filename)
    if not resp.ok and folder_id != "root":
        resp = _try_put("root")
        if _handle_errors(resp):
            return generate_excel(entries, upload_template_bytes, target_filename)

    if not resp.ok:
        logger.error("generate_excel PUT failed: %d %s", resp.status_code, resp.text[:200])

    return resp.ok, resp.json().get("id") if resp.ok else None


def list_folder_files():
    drive_base = _get_drive_base()
    folder_id = _ensure_documents_folder(drive_base=drive_base)
    if not folder_id:
        return []
    resp = requests.get(
        f"{GRAPH_BASE}{drive_base}/items/{folder_id}/children",
        headers=_headers(),
    )
    if not resp.ok:
        if _handle_errors(resp):
            return list_folder_files()
        return []
    files = []
    for child in resp.json().get("value", []):
        name = child.get("name", "")
        is_folder = child.get("folder") is not None
        item = {
            "id": child.get("id"),
            "name": name,
            "size": child.get("size", 0),
            "modified": child.get("lastModifiedDateTime", ""),
            "folder": is_folder,
        }
        if name.lower().endswith((".xlsm", ".xlsx")):
            item["type"] = "file"
            files.append(item)
    files.sort(key=lambda f: f["modified"], reverse=True)
    return files


def _download_file_content(folder_id, filename, drive_base=None):
    if drive_base is None:
        drive_base = _get_drive_base()
    resp = requests.get(
        f"{GRAPH_BASE}{drive_base}/items/{folder_id}:/{filename}:/content",
        headers=_headers(),
    )
    if resp.ok:
        return resp.content
    if _handle_errors(resp):
        return _download_file_content(folder_id, filename, drive_base)
    return None


def _run_drive_op(entries, filename, drive_base):
    """Execute save-to-file against a given drive base. Returns (ok, detail)."""
    folder_id = _ensure_documents_folder(drive_base)
    if not folder_id:
        return False, "Could not find or create the target folder"

    # Save as .xlsx — change extension from .xlsm if needed
    xlsx_name = filename[:-5] + ".xlsx" if filename.lower().endswith(".xlsm") else filename

    # Download the picked file (destination)
    existing = _download_file_content(folder_id, filename, drive_base)
    if existing is None and filename != xlsx_name:
        existing = _download_file_content(folder_id, xlsx_name, drive_base)

    if not existing:
        return False, f"File '{filename}' not found on OneDrive"

    # Try to download the template file for its format (uncorrupted original)
    user = session.get("user", {})
    email = user.get("email", "unknown")
    template_bytes = None
    for ext_is_macro in [True, False]:
        tn = f"template_{_excel_filename(email, ext_is_macro)}"
        tb = _download_file_content(folder_id, tn, drive_base)
        if tb:
            template_bytes = tb
            break

    if template_bytes:
        workbook_bytes, _ = _build_from_template(entries, template_bytes)
    else:
        workbook_bytes, _ = _build_from_template(entries, existing)
    content = workbook_bytes.read()

    upload_headers = _headers()
    upload_headers["Content-Type"] = _XLSX_MIME

    # Delete old .xlsm file if it exists (to avoid stale cached metadata)
    folder_children = requests.get(
        f"{GRAPH_BASE}{drive_base}/items/{folder_id}/children",
        headers=_headers(),
    )
    if folder_children.ok:
        for child in folder_children.json().get("value", []):
            if child.get("name") in (filename, xlsx_name):
                requests.delete(
                    f"{GRAPH_BASE}{drive_base}/items/{child['id']}",
                    headers=_headers(),
                )

    # Create new .xlsx file
    resp = requests.put(
        f"{GRAPH_BASE}{drive_base}/items/{folder_id}:/{xlsx_name}:/content",
        headers=upload_headers, data=content,
    )

    if _handle_errors(resp):
        return _run_drive_op(entries, filename, drive_base)
    if not resp.ok:
        return False, f"Graph API error {resp.status_code}: {resp.text[:300]}"
    return True, resp.json().get("id")


def save_entries_to_file(entries, filename):
    return _run_drive_op(entries, filename, _get_drive_base())


def download_excel():
    user = session.get("user", {})
    email = user.get("email", "unknown")
    drive_base = _get_drive_base()
    folder_id = _ensure_documents_folder(drive_base=drive_base)

    original_name = _get_template_filename()
    filenames_to_try = [original_name] if original_name else []
    for is_macro in (False, True):
        filenames_to_try.append(_excel_filename(email, is_macro))

    for filename in filenames_to_try:
        if not filename:
            continue
        resp = requests.get(
            f"{GRAPH_BASE}{drive_base}/items/{folder_id}:/{filename}:/content",
            headers=_headers(),
        )
        if resp.ok:
            meta_resp = requests.get(
                f"{GRAPH_BASE}{drive_base}/items/{folder_id}:/{filename}",
                headers=_headers(),
            )
            meta = meta_resp.json() if meta_resp.ok else {}
            return resp.content, meta
        if resp.status_code != 404:
            if _handle_errors(resp):
                return download_excel()

    return None, None


def upload_template(file_bytes, filename):
    user = session.get("user", {})
    email = user.get("email", "unknown")
    drive_base = _get_drive_base()
    logger.info("upload_template: drive_base=%s", drive_base)

    folder_id = _ensure_documents_folder(drive_base=drive_base)
    if not folder_id:
        return False, "Could not find or create the target folder"
    logger.info("upload_template: folder_id=%s", folder_id)

    is_macro = filename.lower().endswith(".xlsm") if filename else False
    template_name = f"template_{_excel_filename(email, is_macro)}"

    folder_children = requests.get(
        f"{GRAPH_BASE}{drive_base}/items/{folder_id}/children",
        headers=_headers(),
    )
    existing_id = None
    if folder_children.ok:
        for child in folder_children.json().get("value", []):
            if child.get("name") == template_name:
                existing_id = child["id"]
                break

    if existing_id:
        resp = requests.put(
            f"{GRAPH_BASE}{drive_base}/items/{existing_id}/content",
            headers=_xls_headers(filename), data=file_bytes,
        )
    else:
        resp = requests.put(
            f"{GRAPH_BASE}{drive_base}/items/{folder_id}:/{template_name}:/content",
            headers=_xls_headers(filename), data=file_bytes,
        )

    if _handle_errors(resp):
        return upload_template(file_bytes, filename)

    # Fallback: if upload fails with 404 and folder_id != "root", retry at root level
    if resp.status_code == 404 and folder_id != "root":
        logger.warning("Upload failed with 404, retrying at root level")
        root_resp = requests.put(
            f"{GRAPH_BASE}{drive_base}/items/root:/{template_name}:/content",
            headers=_xls_headers(filename), data=file_bytes,
        )
        if _handle_errors(root_resp):
            return upload_template(file_bytes, filename)
        if root_resp.ok:
            settings_store.set_template_filename(email, filename)
            return True, filename
        return False, f"Graph API error {root_resp.status_code}: {root_resp.text[:200]}"

    if not resp.ok:
        return False, f"Graph API error {resp.status_code}: {resp.text[:200]}"

    settings_store.set_template_filename(email, filename)
    return True, filename


def download_template():
    user = session.get("user", {})
    email = user.get("email", "unknown")
    drive_base = _get_drive_base()
    folder_id = _ensure_documents_folder(drive_base=drive_base)
    logger.info("download_template: drive_base=%s folder_id=%s", drive_base, folder_id)

    for fb_id in (folder_id, "root"):
        if not fb_id:
            continue
        for ext in (".xlsm", ".xlsx"):
            template_name = f"template_{_excel_filename(email, ext == '.xlsm')}"
            resp = requests.get(
                f"{GRAPH_BASE}{drive_base}/items/{fb_id}:/{template_name}:/content",
                headers=_headers(),
            )
            if resp.ok:
                return resp.content
            if resp.status_code != 404:
                if _handle_errors(resp):
                    return download_template()

    return None


def get_document_info():
    try:
        user = session.get("user", {})
        email = user.get("email", "unknown")
        drive_base = _get_drive_base()

        folder_id = _ensure_documents_folder(drive_base=drive_base)
        logger.info("get_document_info: drive_base=%s folder_id=%s", drive_base, folder_id)

        from onedrive_service import load_timesheets
        ts_data = load_timesheets()
        entries = ts_data.get("entries", [])
        entries_count = len(entries)

        info = {"has_document": False, "has_template": False, "document_name": "", "entries_count": entries_count, "folder_id": folder_id, "drive_base": drive_base}

        original_name = _get_template_filename()
        filenames_to_check = [original_name] if original_name else []
        for ext_is_macro, ext in [(False, ".xlsx"), (True, ".xlsm")]:
            fn = _excel_filename(email, ext_is_macro)
            if fn not in filenames_to_check:
                filenames_to_check.append(fn)

        # Check document in folder, then fallback to root
        for filename in filenames_to_check:
            if not filename:
                continue
            for fb_id in (folder_id, "root"):
                resp = requests.get(
                    f"{GRAPH_BASE}{drive_base}/items/{fb_id}:/{filename}",
                    headers=_headers(),
                )
                if resp.ok:
                    data = resp.json()
                    info["has_document"] = True
                    info["document_name"] = data.get("name", filename)
                    info["document_size"] = data.get("size", 0)
                    info["document_modified"] = data.get("lastModifiedDateTime", "")
                    info["document_id"] = data.get("id", "")
                    break
                if resp.status_code == 404:
                    continue
                if _handle_errors(resp):
                    return get_document_info()
            if info["has_document"]:
                break

        # Check template in folder, then fallback to root
        for ext_is_macro, ext in [(False, ".xlsx"), (True, ".xlsm")]:
            template_name = f"template_{_excel_filename(email, ext_is_macro)}"
            for fb_id in (folder_id, "root"):
                t_resp = requests.get(
                    f"{GRAPH_BASE}{drive_base}/items/{fb_id}:/{template_name}",
                    headers=_headers(),
                )
                if t_resp.ok:
                    info["has_template"] = True
                    break
                if t_resp.status_code == 404:
                    continue
                if _handle_errors(t_resp):
                    return get_document_info()
            if info["has_template"]:
                break

        logger.info("get_document_info: result=%s", info)
        return info
    except Exception as e:
        logger.error("get_document_info error: %s", e, exc_info=True)
        return {"has_document": False, "has_template": False, "document_name": "", "entries_count": 0, "error": str(e)}
