import os
import uuid
import io
from datetime import datetime, time
from flask import Blueprint, request, jsonify, session, send_file
from onedrive_service import save_timesheet, load_timesheets, _get_effective_path, _get_access_token as _get_onedrive_token
import onedrive_service
from openpyxl import load_workbook
from document_service import (
    generate_excel,
    download_excel,
    upload_template,
    download_template,
    get_document_info,
    list_folder_files,
    save_entries_to_file,
    _build_workbook,
    _excel_filename,
    _get_template_filename,
    _build_column_map_two_row,
    _build_column_map,
    ALL_HEADER_MAP,
    _has_two_row_header,
)
import gptbots_service
import settings_store
from config import Config

timesheet_bp = Blueprint("timesheet", __name__, url_prefix="/api/timesheet")


def _ensure_authenticated():
    if "user" not in session:
        return False
    return True


def _auto_generate_document():
    try:
        data = load_timesheets()
        entries = data.get("entries", [])
        template_bytes = download_template()
        if not template_bytes:
            return None, "No template found"
        ok, doc_id = generate_excel(entries, template_bytes)
        if not ok:
            return None, f"generate_excel returned False"
        return doc_id, None
    except Exception as e:
        return None, str(e)


# ── Entry CRUD ──────────────────────────────────────────────────


@timesheet_bp.route("/entries", methods=["GET"])
def get_entries():
    if not _ensure_authenticated():
        return jsonify({"error": "Not authenticated"}), 401
    data = load_timesheets()
    return jsonify(data.get("entries", []))


@timesheet_bp.route("/entries", methods=["POST"])
def add_entry():
    if not _ensure_authenticated():
        return jsonify({"error": "Not authenticated"}), 401

    body = request.get_json(silent=True)
    if not body:
        return jsonify({"error": "Invalid JSON"}), 400

    activity_type = body.get("activity_type", "document")
    entry = {
        "id": str(uuid.uuid4()),
        "activity_type": activity_type,
        "date": body.get("date", datetime.utcnow().strftime("%Y-%m-%d")),
        "user_email": session["user"]["email"],
        "user_name": session["user"]["displayName"],
        "created_at": datetime.utcnow().isoformat(),
    }

    if activity_type == "leave_travel":
        entry["leave_travel_type"] = body.get("leave_travel_type", "")
        entry["time"] = float(body.get("time", 8.5) or 8.5)
        entry["project_id"] = ""
        entry["doc_task_type"] = ""
        entry["doc_id"] = ""
        entry["doc_version"] = ""
        entry["doc_type"] = ""
        entry["work_time"] = 0
        entry["reviewer_time"] = 0
        entry["doc_status"] = ""
        entry["activity_code"] = ""
        entry["activity_time"] = 0
        entry["work_location"] = ""
    elif activity_type == "other":
        required = ["activity_code", "activity_time"]
        for field in required:
            if field not in body or not body[field]:
                return jsonify({"error": f"Missing field: {field}"}), 400
        entry["activity_code"] = body["activity_code"]
        entry["activity_time"] = float(body["activity_time"])
        entry["work_location"] = body.get("work_location", "")
        entry["project_id"] = body.get("project_id", "")
        entry["remarks"] = body.get("remarks", "")
        entry["leave_travel_type"] = ""
        entry["time"] = 0
    else:
        required = ["project_id", "doc_task_type", "work_time"]
        for field in required:
            if field not in body or not body[field]:
                return jsonify({"error": f"Missing field: {field}"}), 400
        entry["project_id"] = body["project_id"]
        entry["doc_task_type"] = body["doc_task_type"]
        entry["doc_id"] = body.get("doc_id", "")
        entry["doc_version"] = body.get("doc_version", "")
        entry["doc_type"] = body.get("doc_type", "")
        entry["work_time"] = float(body["work_time"])
        entry["reviewer_time"] = float(body.get("reviewer_time", 0) or 0)
        entry["doc_status"] = body.get("doc_status", "")
        entry["activity_code"] = ""
        entry["activity_time"] = 0
        entry["work_location"] = body.get("work_location", "")
        entry["remarks"] = body.get("remarks", "")
        entry["leave_travel_type"] = ""
        entry["time"] = 0

    data = load_timesheets()
    entries = data.get("entries", [])
    entries.append(entry)

    if not save_timesheet({"entries": entries}):
        return jsonify({"error": "Failed to save entry — check permissions"}), 500
    doc_id, gen_error = _auto_generate_document()
    resp = {"entry": entry, "entries": entries}
    if gen_error:
        resp["generate_warning"] = gen_error
    return jsonify(resp), 201


@timesheet_bp.route("/entries/<entry_id>", methods=["PUT"])
def update_entry(entry_id):
    if not _ensure_authenticated():
        return jsonify({"error": "Not authenticated"}), 401

    body = request.get_json(silent=True)
    if not body:
        return jsonify({"error": "Invalid JSON"}), 400

    data = load_timesheets()
    entries = data.get("entries", [])

    found = False
    for entry in entries:
        if entry.get("id") == entry_id:
            entry["date"] = body.get("date", entry.get("date", ""))
            at = body.get("activity_type", entry.get("activity_type", "document"))
            entry["activity_type"] = at
            if entry.get("activity_type") == "leave_travel":
                entry["leave_travel_type"] = body.get("leave_travel_type", entry.get("leave_travel_type", ""))
                entry["time"] = float(body.get("time", entry.get("time", 8.5)) or 8.5)
                entry["project_id"] = ""
                entry["doc_task_type"] = ""
                entry["doc_id"] = ""
                entry["doc_version"] = ""
                entry["doc_type"] = ""
                entry["work_time"] = 0
                entry["reviewer_time"] = 0
                entry["doc_status"] = ""
                entry["activity_code"] = ""
                entry["activity_time"] = 0
                entry["work_location"] = ""
            elif entry.get("activity_type") == "other":
                entry["activity_code"] = body.get("activity_code", entry.get("activity_code", ""))
                entry["activity_time"] = float(body.get("activity_time", entry.get("activity_time", 0)))
                entry["work_location"] = body.get("work_location", entry.get("work_location", ""))
                entry["project_id"] = body.get("project_id", entry.get("project_id", ""))
                entry["remarks"] = body.get("remarks", entry.get("remarks", ""))
                entry["doc_task_type"] = ""
                entry["doc_id"] = ""
                entry["doc_version"] = ""
                entry["doc_type"] = ""
                entry["work_time"] = 0
                entry["reviewer_time"] = 0
                entry["doc_status"] = ""
                entry["leave_travel_type"] = ""
                entry["time"] = 0
            else:
                entry["project_id"] = body.get("project_id", entry["project_id"])
                entry["doc_task_type"] = body.get("doc_task_type", entry.get("doc_task_type", ""))
                entry["doc_id"] = body.get("doc_id", entry.get("doc_id", ""))
                entry["doc_version"] = body.get("doc_version", entry.get("doc_version", ""))
                entry["doc_type"] = body.get("doc_type", entry.get("doc_type", ""))
                entry["work_time"] = float(body.get("work_time", entry.get("work_time", 0)))
                entry["reviewer_time"] = float(body.get("reviewer_time", entry.get("reviewer_time", 0)) or 0)
                entry["doc_status"] = body.get("doc_status", entry.get("doc_status", ""))
                entry["activity_code"] = ""
                entry["activity_time"] = 0
                entry["work_location"] = body.get("work_location", entry.get("work_location", ""))
                entry["remarks"] = body.get("remarks", entry.get("remarks", ""))
                entry["leave_travel_type"] = ""
                entry["time"] = 0
            entry["user_name"] = session["user"]["displayName"]
            entry["updated_at"] = datetime.utcnow().isoformat()
            found = True
            break

    if not found:
        return jsonify({"error": "Entry not found"}), 404

    if not save_timesheet({"entries": entries}):
        return jsonify({"error": "Failed to save entry — check permissions"}), 500
    _auto_generate_document()
    return jsonify({"entries": entries})


@timesheet_bp.route("/entries/<entry_id>", methods=["DELETE"])
def delete_entry(entry_id):
    if not _ensure_authenticated():
        return jsonify({"error": "Not authenticated"}), 401

    data = load_timesheets()
    entries = data.get("entries", [])
    entries = [e for e in entries if e.get("id") != entry_id]

    if not save_timesheet({"entries": entries}):
        return jsonify({"error": "Failed to save entry — check permissions"}), 500
    _auto_generate_document()
    return jsonify({"entries": entries})


# ── GPTBots Chat ────────────────────────────────────────────────

import json
import re
from datetime import date, timedelta


def _parse_table_from_reply(reply):
    m = re.search(r"^\|(.+)\|\s*$", reply, re.MULTILINE)
    if not m:
        return None
    lines = re.findall(r"^\|(.+)\|\s*$", reply, re.MULTILINE)
    if len(lines) < 2:
        return None
    lines = [l for l in lines if not re.search(r"^[-| :]+$", l.strip())]
    if len(lines) < 1:
        return None
    headers = [h.strip().lower() for h in lines[0].split("|")]
    vals = [v.strip() for v in lines[1].split("|")]
    COLUMN_MAP = {
        "date": "date",
        "type": "activity_type",
        "project_id": "project_id",
        "project id": "project_id",
        "project": "project_id",
        "activity_code": "activity_code",
        "activity code": "activity_code",
        "code": "activity_code",
        "activity_time": "activity_time",
        "activity time": "activity_time",
        "hours": "activity_time",
        "duration": "activity_time",
        "hrs": "activity_time",
        "time": "activity_time",
        "work_location": "work_location",
        "work location": "work_location",
        "location": "work_location",
        "doc_task_type": "doc_task_type",
        "doc task type": "doc_task_type",
        "task_type": "doc_task_type",
        "task type": "doc_task_type",
        "doc_id": "doc_id",
        "doc id": "doc_id",
        "doc_version": "doc_version",
        "doc version": "doc_version",
        "version": "doc_version",
        "doc_type": "doc_type",
        "doc type": "doc_type",
        "work_time": "work_time",
        "work time": "work_time",
        "reviewer_time": "reviewer_time",
        "reviewer time": "reviewer_time",
        "doc_status": "doc_status",
        "doc status": "doc_status",
        "status": "doc_status",
        "remarks": "",
        "remark": "",
        "note": "",
        "notes": "",
    }
    result = {}
    for i, h in enumerate(headers):
        key = COLUMN_MAP.get(h)
        if key is None:
            continue
        if not key:
            continue
        val = vals[i] if i < len(vals) else ""
        if key == "activity_type" and val.lower() in ("meeting", "training", "other", "site visit"):
            val = "other"
        if key == "activity_time":
            try:
                val = float(val)
            except (ValueError, TypeError):
                val = 0.0
        if key == "work_time":
            try:
                val = float(val)
            except (ValueError, TypeError):
                val = 0.0
        result[key] = val
    return result if result else None


def _parse_entry_from_reply(reply):
    raw = None
    m = re.search(r"<entry>(.*?)</entry>", reply, re.DOTALL)
    if m:
        raw = m.group(1).strip()
    if not raw:
        m = re.search(r"```(?:json)?\s*\n(.*?)\n```", reply, re.DOTALL)
        if m:
            raw = m.group(1).strip()
    if not raw:
        m = re.search(r"`(.*?)`", reply)
        if m:
            raw = m.group(1).strip()

    parsed = None
    if raw:
        raw = re.sub(r",\s*}", "}", raw)
        raw = re.sub(r",\s*]", "]", raw)

        for attempt in [raw, raw.replace("'", '"')]:
            try:
                parsed = json.loads(attempt)
                break
            except json.JSONDecodeError:
                pass

        if not parsed:
            try:
                fixed = re.sub(r"(\w+):", r'"\1":', raw)
                fixed = fixed.replace("'", '"')
                parsed = json.loads(fixed)
            except json.JSONDecodeError:
                pass

    table_result = _parse_table_from_reply(reply)
    if parsed and table_result:
        parsed.update(table_result)
        return parsed
    return parsed or table_result


WEEKDAYS = {"monday": 0, "tuesday": 1, "wednesday": 2, "thursday": 3, "friday": 4, "saturday": 5, "sunday": 6}


def _resolve_date(text):
    today = date.today()
    t = text.lower()
    if re.search(r"\b(today|now)\b", t):
        return today.isoformat()
    if re.search(r"\byesterday\b", t):
        return (today - timedelta(days=1)).isoformat()
    if re.search(r"\btomorrow\b", t):
        return (today + timedelta(days=1)).isoformat()

    m_ad = re.search(r"(\d{4}-\d{2}-\d{2})", text)
    if m_ad:
        return m_ad.group(1)

    m_dmy = re.search(r"(\d{1,2})[./-](\d{1,2})[./-](\d{2,4})", text)
    if m_dmy:
        d, mo, y = int(m_dmy.group(1)), int(m_dmy.group(2)), m_dmy.group(3)
        y = int(y) + (2000 if int(y) < 100 else 0)
        try:
            return date(y, mo, d).isoformat()
        except ValueError:
            pass

    for wd_name, wd_num in WEEKDAYS.items():
        m = re.search(rf"(last|this|next)\s+{wd_name}", t)
        if m:
            rel = m.group(1)
            days_ahead = wd_num - today.weekday()
            if rel == "last":
                days_ahead -= 7
            elif rel == "next":
                days_ahead += 7
            return (today + timedelta(days=days_ahead)).isoformat()

    return today.isoformat()


def _resolve_hours(text):
    t = text.lower()
    m = re.search(r"half\s*day", t)
    if m:
        return 4.0
    m = re.search(r"full\s*day", t)
    if m:
        return 8.0

    m = re.search(r"(\d+(?:\.\d+)?)\s*(?:hours?|hrs?|h)\b", text, re.IGNORECASE)
    if m:
        return float(m.group(1))

    m = re.search(r"from\s+(\d{1,2})(?::(\d{2}))?\s*(?:am|a\.m\.)?\s*(?:to|-|until)\s+(\d{1,2})(?::(\d{2}))?\s*(?:pm|p\.m\.)", text, re.IGNORECASE)
    if m:
        start_h, start_m = int(m.group(1)), int(m.group(2) or 0)
        end_h, end_m = int(m.group(3)), int(m.group(4) or 0)
        if end_h < start_h or (end_h == start_h and end_m < start_m):
            end_h += 12
        return round((end_h * 60 + end_m - start_h * 60 - start_m) / 60, 1)

    return 0.0


PROJECT_PATTERN = re.compile(r"\b(\d{4}-[A-Z0-9]+)\b")


def _resolve_project(text):
    m = PROJECT_PATTERN.search(text.upper())
    if m:
        return m.group(1)
    m = re.search(r"(?:project|prj)[:\s]*(\S+)", text, re.IGNORECASE)
    if m:
        return m.group(1).strip().upper()
    return ""


LOCATION_PATTERNS = [
    (r"\b(office\s*5-5)\b", "Office 5-5"),
    (r"\b(office\s*12-7)\b", "Office 12-7"),
    (r"\b(office\s*5-13a)\b", "Office 5-13A"),
    (r"\b(kuching)\b", "Kuching Br."),
    (r"\b(jakarta)\b", "Jakarta Br."),
    (r"\b(bangalore|bangalor)\b", "Bangalore Br."),
    (r"\bsite\b", "Site"),
    (r"\bwfh\b", "WFH Primary"),
    (r"\bhome\b", "WFH Primary"),
    (r"\boffice\s*(\S+)", None),
]


def _resolve_location(text):
    t = text.lower()
    for pattern, replacement in LOCATION_PATTERNS:
        m = re.search(pattern, t)
        if m:
            if replacement:
                return replacement
            return m.group(1).strip().title()
    return ""


ACTIVITY_CODE_MAP = [
    (r"invoicing|invoice", "PC0-Invoicing"),
    (r"p.tracker.*prep|tracker prep", "PC1-P-Tracker Prep-Up"),
    (r"p.tracker.*prog|progress.*tracker", "PC2-P-Tracker-Prog."),
    (r"cash.?flow", "PC3-Cash Flow"),
    (r"contract", "PC4-Contract"),
    (r"closing|project closing", "PC10-Proj. Closing"),
    (r"subcon|subcontractor", "PC5-Subcon"),
    (r"data prep", "PD1-Data Prep"),
    (r"data analysis|data analyze", "PD2-Data Analysis"),
    (r"software dev|development", "PD3-Software Dev"),
    (r"folder.?setup|folder set.?up", "PF1-Folder Set up"),
    (r"archive|docs.*archive", "PF2-Docs. Archive"),
    (r"time.?plan|planning", "PL1-Time Plan"),
    (r"meeting.*client|client.*meeting", "PM1-Mtg. Client"),
    (r"meeting.*internal|internal.*meeting|team.*meeting", "PM2-Mtg. Int. Team"),
    (r"meeting.*(?:cc|consultant|other)", "PM3-Mtg. CC-Others"),
    (r"qc|quality.*check|doc.*check|doc.*review|document.*check|document.*review", "PQ1-Doc. QC"),
    (r"site.?visit|site.?survey", "PS1-Site Visit"),
    (r"test.*site|site.*test(?:ing)?", "PS2-Test Site"),
    (r"test.*factory|factory.*test(?:ing)?", "PS3-Test Factory"),
    (r"d.tracker.*prep|tracker.*prep", "PT1-D-Tracker Prep-Up"),
    (r"d.tracker.*prog|tracker.*progress", "PT2-D-Tracker Prog."),
    (r"trainer|training.*time", "PTR0-Trainer's Time"),
    (r"other.*activity|misc|others", "POA-Others"),
    (r"mtg.*hq|meeting.*hq", "NM0-Mtg. HQ"),
    (r"mtg.*gen|meeting.*general", "NM1-Mtg. Gen"),
    (r"mtg.*fest|meeting.*festival", "NM3-Mtg. Fest."),
    (r"office.?mgmt|office management", "NOM-Office Mgmt."),
    (r"pc.*admin", "NKM-PC-Admin"),
    (r"pc.*recruit|recruitment", "NKM-PC-Recruit"),
    (r"pc.*payroll|payroll", "NKM-PC-Payroll"),
    (r"pc.*statutory|statutory", "NKM-PC-Statutory"),
    (r"pc.*comp.?ben|compensation|benefit", "NKM-PC-CompBen"),
    (r"pc.*er|employee.?relation", "NKM-PC-ER"),
    (r"pc.*contract", "NKM-PC-Contract"),
    (r"pc.*training", "NKM-PC-Training"),
    (r"pc.*talent|talent", "NKM-PC-Talent"),
    (r"pc.*ir|industrial.?relation", "NKM-PC-IR"),
    (r"pc.*policy|policy", "NKM-PC-Policy"),
    (r"pc.*audit", "NKM-PC-Audit"),
    (r"pc.*branding|branding", "NKM-PC-Branding"),
    (r"pc.*wellness|wellness", "NKM-PC-Wellness"),
    (r"pc.*reporting|reporting", "NKM-PC-Reporting"),
    (r"admin.*ops|admin operation", "NKM-Admin-Ops"),
    (r"admin.*facility|facility", "NKM-Admin-Facility"),
    (r"admin.*support|admin support", "NKM-Admin-Support"),
    (r"bd.*mtg.*int|bd.*meeting.*internal", "NKM-BD-Mtg-Int"),
    (r"bd.*mtg.*ext|bd.*meeting.*external", "NKM-BD-Mtg-Ext"),
    (r"bd.*proposal|proposal", "NKM-BD-Proposal"),
    (r"bd.*cv|cv.*prep", "NKM-BD-CV"),
    (r"bd.*tdr|brief|tender.*brief", "NKM-BD-TdrBrief"),
    (r"bd.*doc.?mgmt|bd.*document", "NKM-BD-DocMgmt"),
    (r"bd.*other", "NKM-BD-Other"),
    (r"com.*invoice|commercial.*invoice", "NKM-COM-Invoice"),
    (r"com.*timesheet|commercial.*timesheet", "NKM-COM-Timesheet"),
    (r"com.*vendor|vendor.?reg", "NKM-COM-VendorReg"),
    (r"com.*pay.?mgmt|payment.*mgmt", "NKM-COM-PayMgmt"),
    (r"com.*contract", "NKM-COM-Contract"),
    (r"com.*collection|collection", "NKM-COM-Collection"),
    (r"com.*other", "NKM-COM-Others"),
    (r"fin.*data.?entry|data entry", "NKM-FIN-DataEntry"),
    (r"fin.*einvoice|e.?invoice", "NKM-FIN-Einvoice"),
    (r"fin.*payment|payment", "NKM-FIN-payments"),
    (r"fin.*budget|budget", "NKM-FIN-Budget"),
    (r"fin.*cash.?flow|cashflow", "NKM-FIN-Cashflow"),
    (r"fin.*hq.?report|hq report", "NKM-FIN-HQReport"),
    (r"fin.*month.?close|month.?end", "NKM-FIN-MonthClose"),
    (r"fin.*year.?close|year.?end", "NKM-FIN-YearClose"),
    (r"fin.*tax|tax", "NKM-FIN-Tax"),
    (r"fin.*ye.?audit|year.?end.?audit", "NKM-FIN-YEAudit"),
    (r"fin.*audit.?follow|audit follow.?up", "NKM-FIN-AuditFollow"),
    (r"c.*dev|communication.*development", "NKM-C-Dev"),
    (r"c.*rep|communication.*report", "NKM-C-Rep"),
    (r"c.*mtg|communication.*meeting", "NKM-C-Mtg"),
    (r"c.*awc|award.*ceremony", "NKM-C-Awc"),
    (r"q.*dc|quality.*doc.?control", "NKM-Q-DC"),
    (r"q.*aud|quality.*audit", "NKM-Q-Aud"),
    (r"q.*mtg|quality.*meeting", "NKM-Q-Mtg"),
    (r"q.*rev|quality.*review", "NKM-Q-Rev"),
    (r"rm.*rep|risk.*report", "NKM-RM-Rep"),
    (r"rm.*mtg|risk.*meeting", "NKM-RM-Mtg"),
    (r"trn.*int|training.*internal", "NKM-TRN-Int"),
    (r"trn.*hq|training.*hq", "NKM-TRN-HQ"),
    (r"s.*mtg|strategy.*meeting", "NKM-S-Mtg"),
    (r"s.*dm|strategy.*decision|decision.?making", "NKM-S-DM"),
    (r"s.*dev|strategy.*development", "NKM-S-Dev"),
    (r"s.*rep|strategy.*report", "NKM-S-Rep"),
    (r"g.*mtg|governance.*meeting", "NKM-G-Mtg"),
    (r"g.*com|governance.*communication", "NKM-G-Com"),
    (r"g.*adm|governance.*admin|governance.*administration", "NKM-G-Adm"),
    (r"data.*viz|visualization", "NKM-DATA-VISUALIZATION"),
    (r"data.*research|research", "NKM-DATA-RESEARCH"),
    (r"data.*ai|data.*ml|machine.?learning|artificial.?intelligence", "NKM-DATA-AI&ML"),
    (r"data.*development|data.*dev", "NKM-DATA-DEVELOPMENT"),
    (r"data.*doc|data.*documentation", "NKM-DATA-DOCUMENTATION"),
    (r"data.*internal.*system|internal system", "NKM-DATA-INTERNAL SYSTEM DEVELOPMENT"),
    (r"noa|other.*activity|misc.*nkm", "NOA-Other"),
]


def _resolve_activity_code(text):
    t = text.lower()
    for pattern, code in ACTIVITY_CODE_MAP:
        if re.search(pattern, t):
            return code
    return ""


def _fallback_extract(text):
    t = text.lower()
    hours = _resolve_hours(text)
    if not hours:
        return None

    entry_date = _resolve_date(text)
    project_id = _resolve_project(text)
    work_location = _resolve_location(text)

    is_other = any(w in t for w in ("meeting", "training", "admin", "non-revenue", "activity", "site visit", "leave", "off", "break", "lunch"))
    has_doc_keywords = any(w in t for w in ("document", "doc ", "prepare", "check", "review", "draft", "new doc", "update doc"))

    if is_other and not has_doc_keywords:
        activity_code = _resolve_activity_code(text)
        if not activity_code:
            m_code = re.search(r"(?:code|activity)[:\s]*(\S+)", text, re.IGNORECASE)
            activity_code = m_code.group(1) if m_code else ""
        return {
            "activity_type": "other",
            "date": entry_date,
            "activity_code": activity_code,
            "project_id": project_id,
            "activity_time": hours,
            "work_location": work_location,
        }

    doc_task_type = "P"
    if any(w in t for w in ("check", "review", "verify", "qc", "checked", "reviewed")):
        doc_task_type = "C"

    doc_type = ""
    if any(w in t for w in ("new", "create", "draft", "prepare", "prep", "newly")):
        doc_type = "N"
    elif any(w in t for w in ("update", "revise", "amend", "change", "modify", "revision")):
        doc_type = "U"

    m_doc = re.search(r"(?:doc(?:ument)?)[:\s#]*([A-Za-z0-9][-A-Za-z0-9]*)", text, re.IGNORECASE)
    doc_id = m_doc.group(1).strip() if m_doc else ""

    m_ver = re.search(r"(?:ver(?:sion)?|drs)[:\s#]*([A-Za-z0-9][-A-Za-z0-9.]*)", text, re.IGNORECASE)
    doc_version = m_ver.group(1).strip() if m_ver else ""

    rev_time = 0.0
    m_rev = re.search(r"(?:review(?:er)?|mentor)[^0-9]*(\d+(?:\.\d+)?)\s*(?:h|hours?)", text, re.IGNORECASE)
    if m_rev:
        rev_time = float(m_rev.group(1))

    return {
        "activity_type": "document",
        "date": entry_date,
        "project_id": project_id or "GENERAL",
        "doc_task_type": doc_task_type,
        "doc_id": doc_id,
        "doc_version": doc_version,
        "doc_type": doc_type,
        "work_time": hours,
        "reviewer_time": rev_time,
        "doc_status": "",
        "work_location": work_location,
    }





@timesheet_bp.route("/chat-gptbots", methods=["POST"])
def chat_gptbots():
    if not _ensure_authenticated():
        return jsonify({"error": "Not authenticated"}), 401

    body = request.get_json(silent=True)
    if not body or "message" not in body:
        return jsonify({"error": "Missing 'message' field"}), 400

    msg = body["message"]

    def _detect_activity_type(preview_dict):
        at = preview_dict.get("activity_type", "")
        if at in ("document", "other"):
            return at
        if preview_dict.get("activity_code") or preview_dict.get("activity_time") is not None:
            return "other"
        return "document"

    def _save_preview(preview_dict):
        data = load_timesheets()
        entries = data.get("entries", [])
        at = _detect_activity_type(preview_dict)
        entry = {
            "id": str(uuid.uuid4()),
            "activity_type": at,
            "date": preview_dict.get("date", datetime.utcnow().strftime("%Y-%m-%d")),
            "user_email": session["user"]["email"],
            "user_name": session["user"]["displayName"],
            "created_at": datetime.utcnow().isoformat(),
        }
        if at == "other":
            entry["activity_code"] = preview_dict.get("activity_code", "")
            entry["activity_time"] = float(preview_dict.get("activity_time", preview_dict.get("work_time", 0)) or 0)
            entry["work_location"] = preview_dict.get("work_location", "")
            entry["project_id"] = preview_dict.get("project_id", "")
            entry["remarks"] = preview_dict.get("remarks", "")
            entry["doc_task_type"] = ""
            entry["doc_id"] = ""
            entry["doc_version"] = ""
            entry["doc_type"] = ""
            entry["work_time"] = 0
            entry["reviewer_time"] = 0
            entry["doc_status"] = ""
        else:
            entry["project_id"] = preview_dict.get("project_id", "")
            entry["doc_task_type"] = preview_dict.get("doc_task_type", "")
            entry["doc_id"] = preview_dict.get("doc_id", "")
            entry["doc_version"] = preview_dict.get("doc_version", "")
            entry["doc_type"] = preview_dict.get("doc_type", "")
            entry["work_time"] = float(preview_dict.get("work_time", preview_dict.get("activity_time", 0)) or 0)
            entry["reviewer_time"] = float(preview_dict.get("reviewer_time", 0) or 0)
            entry["doc_status"] = preview_dict.get("doc_status", "")
            entry["activity_code"] = ""
            entry["activity_time"] = 0
            entry["remarks"] = preview_dict.get("remarks", "")
            entry["work_location"] = preview_dict.get("work_location", "")
        entries.append(entry)
        save_timesheet({"entries": entries})
        template_bytes = download_template()
        generate_excel(entries, template_bytes)
        return entry

    # Handle confirmation from frontend
    confirm_entry = body.get("confirm_entry")
    if confirm_entry:
        try:
            entry = _save_preview(confirm_entry)
            return jsonify({"saved": True, "entry": entry})
        except Exception:
            return jsonify({"error": "Failed to save entry"}), 500

    # Handle refinement of an existing draft
    current_draft = body.get("current_draft")
    if current_draft:
        conversation_id = body.get("conversation_id")
        if not conversation_id:
            user_id = session["user"].get("email", session["user"].get("name", "unknown"))
            conversation_id, err = gptbots_service.create_conversation(user_id)
            if err:
                return jsonify({"error": err}), 500

        refine_context = (
            f"The user had this draft:\n"
            f"{json.dumps(current_draft, indent=2)}\n\n"
            f"Their correction: \"{msg}\"\n\n"
            f"Update the draft accordingly."
        )
        result, err = gptbots_service.send_message(conversation_id, msg, refine_context)
        if err:
            return jsonify({"error": err}), 500

        reply = result["reply"]
        entry_preview = _parse_entry_from_reply(reply)

        if not entry_preview:
            entry_preview = current_draft

        if entry_preview:
            entry_preview["date"] = _resolve_date(msg)

        clean_reply = re.sub(r"\s*<entry>.*?</entry>\s*", "", reply, flags=re.DOTALL).strip()
        return jsonify({
            "reply": clean_reply or "Draft updated based on your correction.",
            "conversation_id": result["conversation_id"],
            "entry_preview": entry_preview,
        })

    user_id = session["user"].get("email", session["user"].get("name", "unknown"))
    conversation_id = body.get("conversation_id")

    if not conversation_id:
        conversation_id, err = gptbots_service.create_conversation(user_id)
        if err:
            return jsonify({"error": err}), 500

    entry_context = None
    if body.get("include_context", True):
        try:
            data = load_timesheets()
            entries = data.get("entries", [])
            lines = ["id | date | type | project_id | task_type | doc_id | version | doc_type | work_time | reviewer_time | status | activity_code | activity_time | location | remarks"]
            if entries:
                recent = sorted(entries, key=lambda e: e.get("created_at", ""), reverse=True)[:20]
                for e in recent:
                    at = e.get("activity_type", "document")
                    lines.append(f"{e.get('id','')} | {e.get('date','')} | {at} | {e.get('project_id','')} | {e.get('doc_task_type','')} | {e.get('doc_id','')} | {e.get('doc_version','')} | {e.get('doc_type','')} | {e.get('work_time',0)} | {e.get('reviewer_time',0)} | {e.get('doc_status','')} | {e.get('activity_code','')} | {e.get('activity_time',0)} | {e.get('work_location','')} | {e.get('remarks','')}")
            else:
                lines.append("(no entries yet)")
            entry_context = "Here are the current timesheet entries (recent 20):\n" + "\n".join(lines)
        except PermissionError:
            entry_context = None

    result, err = gptbots_service.send_message(conversation_id, msg, entry_context)
    if err:
        return jsonify({"error": err}), 500

    reply = result["reply"]
    entry_preview = _parse_entry_from_reply(reply)

    if entry_preview:
        entry_preview["date"] = _resolve_date(msg)

    clean_reply = re.sub(r"\s*<entry>.*?</entry>\s*", "", reply, flags=re.DOTALL).strip()
    clean_reply = re.sub(r"```(?:json)?\s*\n(.*?)```\s*", "", clean_reply, flags=re.DOTALL).strip()

    return jsonify({
        "reply": clean_reply,
        "conversation_id": result["conversation_id"],
        "message_id": result.get("message_id"),
        "entry_preview": entry_preview,
    })


# ── Document Management ─────────────────────────────────────────


@timesheet_bp.route("/document/upload", methods=["POST"])
def upload_document_template():
    if not _ensure_authenticated():
        return jsonify({"error": "Not authenticated"}), 401

    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files["file"]
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        return jsonify({"error": "Only .xlsx or .xlsm files are accepted as templates"}), 400

    file_bytes = file.read()
    ok, err = upload_template(file_bytes, file.filename)
    if not ok:
        return jsonify({"error": f"Failed to upload template: {err}"}), 500

    data = load_timesheets()
    entries = data.get("entries", [])
    generate_excel(entries, file_bytes)

    return jsonify({"message": "Template uploaded and document regenerated"})


@timesheet_bp.route("/document/download", methods=["GET"])
def download_document():
    if not _ensure_authenticated():
        return jsonify({"error": "Not authenticated"}), 401

    data = load_timesheets()
    entries = data.get("entries", [])
    template_bytes = download_template()
    workbook_bytes, is_macro = _build_workbook(entries, template_bytes)
    content = workbook_bytes.read()

    # Upload to OneDrive in background so next download can work without entries
    try:
        generate_excel(entries, template_bytes)
    except Exception:
        pass

    dl_name = _get_template_filename() or _excel_filename(session["user"]["email"], is_macro)
    # Ensure extension matches actual content
    base, _ = os.path.splitext(dl_name)
    dl_ext = ".xlsm" if is_macro else ".xlsx"
    dl_name = base + dl_ext
    mimetype = "application/vnd.ms-excel.sheet.macroEnabled.12" if is_macro else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return send_file(
        io.BytesIO(content),
        mimetype=mimetype,
        as_attachment=True,
        download_name=dl_name,
    )


@timesheet_bp.route("/document/preview", methods=["GET"])
def preview_document():
    if not _ensure_authenticated():
        return jsonify({"error": "Not authenticated"}), 401
    try:
        def _serialize(v):
            if v is None:
                return None
            if isinstance(v, (datetime, time)):
                return v.isoformat()
            return v

        data = load_timesheets()
        entries = data.get("entries", [])
        template_bytes = download_template()
        workbook_bytes, _ = _build_workbook(entries, template_bytes)
        if not workbook_bytes:
            return jsonify({"error": "Failed to build workbook"}), 500

        wb = load_workbook(workbook_bytes, data_only=True)
        ws = wb.active

        from document_service import _has_two_row_header
        if _has_two_row_header(ws):
            headers = [_serialize(cell.value) for cell in ws[2]]
            rows = []
            for row in ws.iter_rows(min_row=3, values_only=True):
                if any(cell is not None for cell in row):
                    rows.append([_serialize(v) for v in row])
        else:
            headers = [_serialize(cell.value) for cell in ws[1]]
            rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(cell is not None for cell in row):
                    rows.append([_serialize(v) for v in row])

        return jsonify({"headers": headers, "rows": rows})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@timesheet_bp.route("/document/info", methods=["GET"])
def document_info():
    if not _ensure_authenticated():
        return jsonify({"error": "Not authenticated"}), 401
    info = get_document_info()
    data = load_timesheets()
    entries = data.get("entries", [])
    info["entries_count"] = len(entries)
    return jsonify(info)


@timesheet_bp.route("/document/generate", methods=["POST"])
def generate_document():
    if not _ensure_authenticated():
        return jsonify({"error": "Not authenticated"}), 401

    data = load_timesheets()
    entries = data.get("entries", [])
    if not entries:
        return jsonify({"error": "No entries to write"}), 400

    template_bytes = download_template()
    ok, doc_id = generate_excel(entries, template_bytes)
    if not ok:
        return jsonify({"error": "Failed to generate document"}), 500
    return jsonify({"message": "Document generated successfully", "file_id": doc_id})


@timesheet_bp.route("/document/generate-by-month", methods=["POST"])
def generate_document_by_month():
    if not _ensure_authenticated():
        return jsonify({"error": "Not authenticated"}), 401

    body = request.get_json(silent=True)
    if not body or "year" not in body or "month" not in body:
        return jsonify({"error": "Missing 'year' or 'month' fields"}), 400

    year = int(body["year"])
    month = int(body["month"])
    if month < 1 or month > 12:
        return jsonify({"error": "Month must be 1-12"}), 400

    data = load_timesheets()
    all_entries = data.get("entries", [])

    filtered = []
    for e in all_entries:
        try:
            d = datetime.strptime(e.get("date", ""), "%Y-%m-%d")
            if d.year == year and d.month == month:
                filtered.append(e)
        except (ValueError, TypeError):
            continue

    if not filtered:
        return jsonify({"error": f"No entries found for {year}-{month:02d}"}), 400

    template_bytes = download_template()
    ok, doc_id = generate_excel(filtered, template_bytes)
    if not ok:
        return jsonify({"error": "Failed to generate document"}), 500
    return jsonify({"message": f"Document generated for {year}-{month:02d}", "file_id": doc_id, "entry_count": len(filtered)})


@timesheet_bp.route("/documents/list", methods=["GET"])
def list_documents():
    if not _ensure_authenticated():
        return jsonify({"error": "Not authenticated"}), 401
    effective_path = _get_effective_path()
    files = list_folder_files()
    return jsonify({
        "files": files,
        "path": effective_path,
        "count": len(files),
    })


@timesheet_bp.route("/documents/save-to-file", methods=["POST"])
def save_to_file():
    if not _ensure_authenticated():
        return jsonify({"error": "Not authenticated"}), 401
    body = request.get_json(silent=True)
    if not body or "filename" not in body:
        return jsonify({"error": "Missing 'filename' field"}), 400
    filename = body["filename"].strip()
    if not filename:
        return jsonify({"error": "Filename cannot be empty"}), 400
    data = load_timesheets()
    entries = data.get("entries", [])
    ok, detail = save_entries_to_file(entries, filename)
    if not ok:
        return jsonify({"error": detail or "Failed to save to file"}), 500
    return jsonify({"message": f"Entries written to {filename}", "file_id": detail})


@timesheet_bp.route("/debug/drive", methods=["GET"])
def debug_drive():
    if not _ensure_authenticated():
        return jsonify({"error": "Not authenticated"}), 401
    from onedrive_service import _get_drive_base, GRAPH_BASE, _get_effective_path, _ensure_timesheets_folder
    import requests as req
    token, _ = onedrive_service._get_access_token()
    drives_resp = req.get(f"{GRAPH_BASE}/me/drives", headers={"Authorization": f"Bearer {token}"})
    drives_data = drives_resp.json() if drives_resp.ok else drives_resp.text
    shared_resp = req.get(f"{GRAPH_BASE}/me/drive/sharedWithMe", headers={"Authorization": f"Bearer {token}"})
    shared_data = shared_resp.json() if shared_resp.ok else shared_resp.text
    recent_resp = req.get(f"{GRAPH_BASE}/me/drive/recent", headers={"Authorization": f"Bearer {token}"})
    recent_data = recent_resp.json() if recent_resp.ok else recent_resp.text
    drive_base = _get_drive_base()
    return jsonify({
        "drive_base": drive_base,
        "drives_status": drives_resp.status_code,
        "drives": drives_data,
        "shared_status": shared_resp.status_code,
        "shared": shared_data,
        "recent_status": recent_resp.status_code,
        "recent": recent_data,
        "sharepoint_host": Config.SHAREPOINT_SITE_HOST,
        "sharepoint_path": Config.SHAREPOINT_SITE_PATH,
        "onedrive_root": Config.ONEDRIVE_ROOT_PATH,
    })


@timesheet_bp.route("/debug/entries-storage", methods=["GET"])
def debug_entries_storage():
    if not _ensure_authenticated():
        return jsonify({"error": "Not authenticated"}), 401
    from onedrive_service import _get_drive_base, GRAPH_BASE, _get_effective_path, _ensure_timesheets_folder, _get_access_token
    import requests as req
    email = session["user"]["email"]
    effective_path = _get_effective_path()
    drive_base = _get_drive_base()
    folder_id = _ensure_timesheets_folder()
    filename = f"timesheet_{email.replace('@', '_at_')}.json"
    result = {
        "email": email,
        "effective_path": effective_path,
        "drive_base": drive_base,
        "folder_id": folder_id,
        "filename": filename,
        "folder_children": [],
        "json_content": None,
        "json_status": None,
        "error": None,
    }
    if folder_id:
        try:
            token, _ = _get_access_token()
            children = req.get(
                f"{GRAPH_BASE}{drive_base}/items/{folder_id}/children",
                headers={"Authorization": f"Bearer {token}"},
            )
            if children.ok:
                kids = []
                for c in children.json().get("value", []):
                    kids.append({"name": c.get("name"), "id": c.get("id")})
                result["folder_children"] = kids
        except Exception as e:
            result["error"] = str(e)
        try:
            token, _ = _get_access_token()
            json_resp = req.get(
                f"{GRAPH_BASE}{drive_base}/items/{folder_id}:/{filename}:/content",
                headers={"Authorization": f"Bearer {token}"},
            )
            result["json_status"] = json_resp.status_code
            if json_resp.ok:
                data = json_resp.json()
                result["json_content"] = {
                    "entry_count": len(data.get("entries", [])),
                    "keys": list(data.keys()),
                }
        except Exception as e:
            result["error"] = str(e)
    return jsonify(result)


@timesheet_bp.route("/debug/token-scopes", methods=["GET"])
def debug_token_scopes():
    if not _ensure_authenticated():
        return jsonify({"error": "Not authenticated"}), 401
    import base64, json
    try:
        token, _ = onedrive_service._get_access_token()
        parts = token.split(".")
        padded = parts[1] + "=" * (4 - len(parts[1]) % 4)
        payload = base64.urlsafe_b64decode(padded)
        claims = json.loads(payload)
        return jsonify({
            "scp": claims.get("scp", "none"),
            "aud": claims.get("aud", ""),
            "iss": claims.get("iss", ""),
            "exp": claims.get("exp", ""),
            "roles": claims.get("roles", []),
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Settings ────────────────────────────────────────────────────


@timesheet_bp.route("/settings/path", methods=["GET"])
def get_settings_path():
    if not _ensure_authenticated():
        return jsonify({"error": "Not authenticated"}), 401
    email = session["user"]["email"]
    custom = settings_store.get_timesheet_path(email)
    return jsonify({
        "path": custom or Config.ONEDRIVE_ROOT_PATH,
        "is_custom": bool(custom),
    })


@timesheet_bp.route("/settings/path", methods=["PUT"])
def set_settings_path():
    if not _ensure_authenticated():
        return jsonify({"error": "Not authenticated"}), 401
    body = request.get_json(silent=True)
    if not body or "path" not in body:
        return jsonify({"error": "Missing 'path' field"}), 400
    path = body["path"].strip()
    if not path:
        return jsonify({"error": "Path cannot be empty"}), 400
    email = session["user"]["email"]
    settings_store.set_timesheet_path(email, path)
    return jsonify({"path": path, "is_custom": True})


@timesheet_bp.route("/debug-column-mapping", methods=["GET"])
def debug_column_mapping():
    if not _ensure_authenticated():
        return jsonify({"error": "Not authenticated"}), 401
    try:
        import io
        data = load_timesheets()
        entries = data.get("entries", [])
        template_bytes = download_template()
        if not template_bytes:
            return jsonify({"error": "No template found"}), 404

        wb = load_workbook(io.BytesIO(template_bytes))
        ws = wb.active
        is_2row = _has_two_row_header(ws)
        header_row = 2 if is_2row else 1

        col_map = _build_column_map_two_row(ws, ALL_HEADER_MAP) if is_2row else _build_column_map(ws, ALL_HEADER_MAP)
        if not is_2row:
            col_map = {k: [v] for k, v in col_map.items()}

        col_map_str = {}
        for field, cols in sorted(col_map.items()):
            header_vals = []
            for c in cols:
                cell_val = str(ws.cell(row=header_row, column=c).value or "")[:40]
                header_vals.append({"col": c, "header": cell_val})
            col_map_str[field] = header_vals

        act_fields = ["activity_code", "project_id", "activity_time"]
        slot_info = []
        for slot_i in range(3):
            slot = {"slot": slot_i, "fields": []}
            for fi, field in enumerate(act_fields):
                if field == "project_id":
                    idx = 2 + slot_i
                else:
                    idx = slot_i
                lst = col_map.get(field, [])
                col = lst[idx] if idx < len(lst) else None
                expected = 19 + slot_i * 3 + fi
                slot["fields"].append({
                    "field": field,
                    "col_detected": col,
                    "col_expected": expected,
                    "match": col == expected,
                })
            slot_info.append(slot)

        other_entries = [e for e in entries if e.get("activity_type") == "other"]
        sample = []
        for e in other_entries[:5]:
            sample.append({
                "date": e.get("date"),
                "activity_code": e.get("activity_code"),
                "project_id": e.get("project_id"),
                "activity_time": e.get("activity_time"),
            })

        def get_cell(r, c):
            return str(ws.cell(row=r, column=c).value or "")

        header_vals = [get_cell(header_row, c) for c in range(19, 28)]
        data_rows = []
        for r in range(header_row + 1, ws.max_row + 1):
            date_val = get_cell(r, 1)
            vals = [get_cell(r, c) for c in range(19, 28)]
            if any(v for v in vals):
                data_rows.append({"date": date_val, "cols_19_27": vals})

        return jsonify({
            "is_2row_header": is_2row,
            "col_map": col_map_str,
            "slot_mapping": slot_info,
            "sample_other_entries": sample,
            "template_headers_19_27": header_vals,
            "data_rows": data_rows,
        })
    except Exception as e:
        import traceback
        return jsonify({"error": str(e), "traceback": traceback.format_exc()}), 500
